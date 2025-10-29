# ----------------------------------------
# Name    : Audrey
# Note    : This is for creating a xlsx file for listing a bunch of files in a single folder by last write time.
#
# Date Created : July 15, 2025 @8:50PM
# ----------------------------------------

# public library imports
import datetime
import sys
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_points
from PIL import Image as PILImage

dictionary_file1 = 'name_of_file.txt'
dictionary_file2 = 'last_writetime.txt'

class Excel_Filelist:
    """
    A class to handle file operations and Excel spreadsheet management.
    """
    def __init__(self, worksheet=None, dictionary_file1='name_of_file.txt', dictionary_file2='last_writetime.txt'):
        self.ws = worksheet
        self.dictionary_file1 = dictionary_file1
        self.dictionary_file2 = dictionary_file2
        self.dictionaries = {}

    def import_dictionary(self, filename1, filename2):
        dictionary = {}

        try:
            f1 = open(filename1, 'r', encoding='utf-16')
            f2 = open(filename2, 'r', encoding='utf-16')
            list1 = f1.readlines()
            list2 = f2.readlines()
            for item1, item2 in zip(list1, list2):
                item1 = item1.strip('\n')
                item2 = item2.strip('\n')
                dictionary.update({item1: item2})

        except:
            print('list_of_files.txt file does not exist')
        return dictionary

    def set_column_width_pixels(self, col_letter, width=8.43):
        """
        Set column width using pixel measurement.

        Parameters:
        col_letter: column letter (e.g., 'A', 'B', 'C')
            col_letter = 'A'
        width = 8.43 # default column length in xlsx sheets?
        """
        ws.column_dimensions[col_letter].width = 2 * width
        return ws.column_dimensions[col_letter].width


    def print_dictionary(self, dictionary):
        for item1, item2 in dictionary.items():
            print(f"\t{item1}\t{item2}")

    # image = 2025-02-07 144957 switch_before_go.png
    # image_dir = C:\Users\Audrey\OneDrive\Pictures\screenshot-collages
    # saved_image_dir = C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100
    def filelist_thumbnail(self, iter=0, image=None,
                           image_dir=r"C:\Users\Audrey\OneDrive\Pictures\screenshot-collages",
                           saved_image_dir=r"C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100"):

        # Add image to my filelist spreadsheet
        image_path = rf"{image_dir}\{image}"
        cell_address = f"A{iter+3}"

        # Load with Pillow first to get pixel dimensions
        pil_img = PILImage.open(image_path)
        width_px, height_px = pil_img.size
        # print(f"Original Pixel dimensions: {width_px} x {height_px}")

        # Convert Excel column width → pixels (approximation)
        scale = 100
        # Scale height proportionally, height/width
        proportional_factor = height_px / width_px
        target_height_px = int(proportional_factor * scale)
        # print(f"New Pixel dimensions: {scale} x {target_height_px}")

        # (width, height) makes image proportional to uniform width for xlsx sheet.
        pil_img = pil_img.resize((scale, target_height_px))
        saved_image_path = rf"{saved_image_dir}\{image}"
        pil_img.save(saved_image_path)
        pil_img.close()  # IMPORTANT: Close PIL image before openpyxl reads it

        # Load into openpyxl and anchor
        img = Image(saved_image_path)

        img.anchor = cell_address
        ws.row_dimensions[iter+3].height = pixels_to_points(target_height_px)
        ws.add_image(img)
        filelist_wb.save(self.ws)
        return saved_image_path

def add_to_spreadsheet(dictionary, letter1, letter2):
    return_str = ""
    i = 3
    for key, value in dictionary.items():
        name_cell1 = f"{letter1}{i}"
        name_cell2 = f"{letter2}{i}"
        ws[name_cell1] = f"{key}"
        ws[name_cell2] = f"{value}"
        return_str = return_str + f"{name_cell1}\t{key}\t\t\t{name_cell2}\t{value}\n"
        i += 1
    return return_str

if __name__ == '__main__':

    print(sys.executable)
    print(sys.path)

    dt = datetime.datetime.today()
    date = dt.strftime("%Y %B %d @%I:%M%p")

    filelist_wb = Workbook()
    ws = filelist_wb.active
    ws.title = "List of Files"

    ws['B1'] = "Filelist of folder"
    ws['C1'] = date
    print(date)
    #"2025-07-16 11:04AM"
    ws['A2'] = "Image"
    ws['B2'] = "Name"
    ws['C2'] = "LastWriteTime"

    excel_fl = Excel_Filelist('filelist.xlsx', 'name_of_file.txt', 'last_writetime.txt')

    width = 10
    excel_fl.set_column_width_pixels('A', width)
    excel_fl.set_column_width_pixels('B', width*2)
    excel_fl.set_column_width_pixels('C', width*1.5)

    i = 3
    dictionary = excel_fl.import_dictionary(dictionary_file1, dictionary_file2)
    imgname1, timestamp2 = next(iter(dictionary.items()))

    name_cell = f"B{i}"
    lastwritetime_cell = f"C{i}"

    ws[name_cell] = f"{imgname1}"
    ws[lastwritetime_cell] = f"{timestamp2}"

    # excel_fl.print_dictionary(dictionary1)
    # print("\n")
    # excel_fl.print_dictionary(dictionary2)

    print("--------------------\n")

    image_ex = r"2025-02-07 144957 switch_before_go.png"
    image_dir = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-collages"
    saved_image_dir = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100"

    # i is the counter for number of dictionary items. The (image1, timestamp1) unpacks the dictionary items.
    for i, (image1, timestamp1) in enumerate(dictionary.items()):
        if i >= 10:
            break
        returned_path = excel_fl.filelist_thumbnail(iter=i, image=image1, image_dir=image_dir, saved_image_dir=saved_image_dir)
        print(returned_path)

    add_to_spreadsheet(dictionary, "B", "C")

    filelist_wb.save('filelist.xlsx')
    print(f"\n{ws.title} spreadsheet is saved!")